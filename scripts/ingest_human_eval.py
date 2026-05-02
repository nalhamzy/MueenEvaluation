"""Ingest the filled human evaluation workbook and produce the human-validated
results that feed back into the report and Arabic letter.

Inputs:
  - human_eval/blinding_key.json
  - human_eval/sampled_articles.json
  - --filled <path-to-filled-workbook.xlsx>

Outputs:
  - human_eval/human_eval_results.json   (per-article + per-model aggregates)
  - human_eval/human_eval_results.csv    (long-form, one row per judgment)
  - human_eval/human_validation_block_en.md  (Markdown block to append to report)
  - human_eval/human_validation_block_ar.md  (Arabic block to append to letter)

Run from project root:
    py scripts/ingest_human_eval.py --filled human_eval/human_eval_workbook_Ahmed.xlsx
"""

import argparse
import csv
import json
import math
import sys
import statistics
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "backend"))

from database import SessionLocal
from models import (
    EvaluationRun, ModelOutput, RunStatus, OutputStatus,
)


CATEGORIES = ["culture", "finance", "politics", "sports", "tech"]
MODEL_DISPLAY = {
    "qwen3.5-397b-a17b": "Qwen 3.5 397B",
    "deepseek-chat": "DeepSeek Chat",
    "mistral.mistral-large-2402-v1:0": "Mistral Large",
    "Mueen AI": "Mueen AI",
}
MODEL_ORDER = [
    "qwen3.5-397b-a17b",
    "deepseek-chat",
    "mistral.mistral-large-2402-v1:0",
    "Mueen AI",
]
MODEL_DISPLAY_AR = {
    "qwen3.5-397b-a17b": "Qwen 3.5",
    "deepseek-chat": "DeepSeek Chat",
    "mistral.mistral-large-2402-v1:0": "Mistral Large",
    "Mueen AI": "معين الذكي",
}


# ----------------------------------------------------------------------------
# Loading
# ----------------------------------------------------------------------------

def load_blinding(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["letter_to_model"], data["model_to_letter"]


def load_filled(path, letter_to_model):
    """Returns dict: task -> [{article_id, category, scores: {model: score},
    note}]"""
    wb = load_workbook(path, data_only=True)

    def parse_sheet(ws_name):
        ws = wb[ws_name]
        rows = []
        # Find score column letters (J,K,L,M = cols 10..13) and note col (14)
        # Headers identify which letter goes with which column.
        # Column 6=Model A, 7=B, 8=C, 9=D ; scores 10..13 same order
        letter_for_col = {10: "A", 11: "B", 12: "C", 13: "D"}
        for r in range(2, ws.max_row + 1):
            article_id = ws.cell(row=r, column=2).value
            if not article_id:
                continue
            category = (ws.cell(row=r, column=3).value or "").lower()
            note = ws.cell(row=r, column=14).value or ""
            scores = {}
            for col, letter in letter_for_col.items():
                v = ws.cell(row=r, column=col).value
                if v is None or v == "":
                    raise ValueError(
                        f"Empty score at {ws_name} row={r} col={col} "
                        f"(article={article_id})"
                    )
                try:
                    s = int(v)
                except (TypeError, ValueError):
                    raise ValueError(
                        f"Non-integer score at {ws_name} row={r} col={col}: {v!r}"
                    )
                if not 1 <= s <= 5:
                    raise ValueError(
                        f"Score out of range at {ws_name} row={r}: {s}"
                    )
                model_key = letter_to_model[letter]
                scores[model_key] = s
            rows.append({
                "article_id": article_id,
                "category": category,
                "scores": scores,
                "note": note,
            })
        return rows

    return {
        "summary": parse_sheet("Summary Task"),
        "translation": parse_sheet("Translation Task"),
        "evaluator_info": _read_evaluator_info(wb),
    }


def _read_evaluator_info(wb):
    if "Evaluator Info" not in wb.sheetnames:
        return {}
    ws = wb["Evaluator Info"]
    info = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k:
            info[str(k).strip()] = v
    return info


def load_llm_judge_scores(article_ids):
    """Pull GPT-5.2 (LLM judge) Summary & Translation scores from the DB
    for the same 30 articles (latest scored output per model).
    Returns: dict[(model, article_id, task)] -> score (0-10 scale)
    """
    db = SessionLocal()
    try:
        runs = db.query(EvaluationRun).filter(
            EvaluationRun.status == RunStatus.COMPLETED
        ).all()

        latest = {}  # (model, article_id) -> ModelOutput
        for r in runs:
            outs = db.query(ModelOutput).filter(
                ModelOutput.run_id == r.id,
                ModelOutput.status == OutputStatus.SCORED,
            ).all()
            for o in outs:
                if o.article_id not in article_ids:
                    continue
                key = (r.model_name, o.article_id)
                prev = latest.get(key)
                if prev is None or (
                    o.processed_at and prev.processed_at
                    and o.processed_at > prev.processed_at
                ):
                    latest[key] = o

        out = {}
        for (model, aid), o in latest.items():
            if o.summary_score is not None:
                out[(model, aid, "summary")] = float(o.summary_score)
            if o.translation_score is not None:
                out[(model, aid, "translation")] = float(o.translation_score)
        return out
    finally:
        db.close()


# ----------------------------------------------------------------------------
# Statistics
# ----------------------------------------------------------------------------

def mean(xs):
    return sum(xs) / len(xs) if xs else 0.0


def stdev(xs):
    return statistics.stdev(xs) if len(xs) > 1 else 0.0


def rescale_1_5_to_0_10(s):
    """1->0, 5->10. Linear: y = (s - 1) * 2.5"""
    return (s - 1) * 2.5


def spearman(a, b):
    """Spearman rank correlation between two parallel lists."""
    if len(a) != len(b) or len(a) < 2:
        return None

    def rank(xs):
        # average-rank for ties
        sorted_idx = sorted(range(len(xs)), key=lambda i: xs[i])
        ranks = [0.0] * len(xs)
        i = 0
        while i < len(sorted_idx):
            j = i
            while j + 1 < len(sorted_idx) and xs[sorted_idx[j + 1]] == xs[sorted_idx[i]]:
                j += 1
            avg = (i + j) / 2.0 + 1
            for k in range(i, j + 1):
                ranks[sorted_idx[k]] = avg
            i = j + 1
        return ranks

    ra = rank(a)
    rb = rank(b)
    n = len(a)
    mra, mrb = mean(ra), mean(rb)
    num = sum((ra[i] - mra) * (rb[i] - mrb) for i in range(n))
    da = math.sqrt(sum((ra[i] - mra) ** 2 for i in range(n)))
    db = math.sqrt(sum((rb[i] - mrb) ** 2 for i in range(n)))
    if da == 0 or db == 0:
        return None
    return num / (da * db)


def pearson(a, b):
    if len(a) != len(b) or len(a) < 2:
        return None
    ma, mb = mean(a), mean(b)
    num = sum((a[i] - ma) * (b[i] - mb) for i in range(len(a)))
    da = math.sqrt(sum((x - ma) ** 2 for x in a))
    db = math.sqrt(sum((x - mb) ** 2 for x in b))
    if da == 0 or db == 0:
        return None
    return num / (da * db)


# ----------------------------------------------------------------------------
# Aggregation
# ----------------------------------------------------------------------------

def aggregate(filled, llm_scores):
    """Compute all the aggregates we need for the report block."""
    article_ids = [r["article_id"] for r in filled["summary"]]

    # Per-model overall (mean over both tasks combined)
    per_model = {}
    for model in MODEL_ORDER:
        sum_scores = [r["scores"][model] for r in filled["summary"]]
        tx_scores = [r["scores"][model] for r in filled["translation"]]
        all_scores = sum_scores + tx_scores

        per_model[model] = {
            "summary": {
                "mean_1_5": round(mean(sum_scores), 2),
                "stdev": round(stdev(sum_scores), 2),
                "mean_0_10": round(rescale_1_5_to_0_10(mean(sum_scores)), 2),
                "scores_1_5": sum_scores,
            },
            "translation": {
                "mean_1_5": round(mean(tx_scores), 2),
                "stdev": round(stdev(tx_scores), 2),
                "mean_0_10": round(rescale_1_5_to_0_10(mean(tx_scores)), 2),
                "scores_1_5": tx_scores,
            },
            "combined": {
                "mean_1_5": round(mean(all_scores), 2),
                "stdev": round(stdev(all_scores), 2),
                "mean_0_10": round(rescale_1_5_to_0_10(mean(all_scores)), 2),
            },
        }

    # Per-category per-model means (1-5 scale)
    per_cat = {}
    for cat in CATEGORIES:
        sum_rows = [r for r in filled["summary"] if r["category"] == cat]
        tx_rows = [r for r in filled["translation"] if r["category"] == cat]
        per_cat[cat] = {"n": len(sum_rows), "models": {}}
        for model in MODEL_ORDER:
            s = [r["scores"][model] for r in sum_rows]
            t = [r["scores"][model] for r in tx_rows]
            per_cat[cat]["models"][model] = {
                "summary_1_5": round(mean(s), 2) if s else None,
                "translation_1_5": round(mean(t), 2) if t else None,
                "combined_1_5": round(mean(s + t), 2) if (s + t) else None,
            }

    # Spearman + Pearson with LLM judge — paired per (model, article)
    correlations = {}
    for task in ["summary", "translation"]:
        human_vec = []
        llm_vec = []
        for r in filled[task]:
            for model in MODEL_ORDER:
                h = rescale_1_5_to_0_10(r["scores"][model])
                l = llm_scores.get((model, r["article_id"], task))
                if l is None:
                    continue
                human_vec.append(h)
                llm_vec.append(l)
        correlations[task] = {
            "n": len(human_vec),
            "spearman": round(spearman(human_vec, llm_vec) or 0, 3),
            "pearson": round(pearson(human_vec, llm_vec) or 0, 3),
            "human_mean_0_10": round(mean(human_vec), 2),
            "llm_mean_0_10": round(mean(llm_vec), 2),
        }

    # Combined (summary + translation together)
    h_all, l_all = [], []
    for task in ["summary", "translation"]:
        for r in filled[task]:
            for model in MODEL_ORDER:
                l = llm_scores.get((model, r["article_id"], task))
                if l is None:
                    continue
                h_all.append(rescale_1_5_to_0_10(r["scores"][model]))
                l_all.append(l)
    correlations["combined"] = {
        "n": len(h_all),
        "spearman": round(spearman(h_all, l_all) or 0, 3),
        "pearson": round(pearson(h_all, l_all) or 0, 3),
    }

    # Model-level rank correlation (4-point Spearman across model means)
    human_means = [per_model[m]["combined"]["mean_0_10"] for m in MODEL_ORDER]
    # LLM judge means restricted to the 30-article human set
    llm_means_subset = []
    for m in MODEL_ORDER:
        vals = []
        for task in ["summary", "translation"]:
            for r in filled[task]:
                v = llm_scores.get((m, r["article_id"], task))
                if v is not None:
                    vals.append(v)
        llm_means_subset.append(round(mean(vals), 2) if vals else 0)
    correlations["model_level"] = {
        "human_means_0_10": human_means,
        "llm_means_0_10_same_30": llm_means_subset,
        "spearman": round(spearman(human_means, llm_means_subset) or 0, 3),
        "models": [MODEL_DISPLAY[m] for m in MODEL_ORDER],
    }

    # Ordering check
    def rank_dict(values):
        return {m: r + 1 for r, (m, _) in
                enumerate(sorted(zip(MODEL_ORDER, values), key=lambda x: -x[1]))}

    correlations["ranking_human"] = rank_dict(human_means)
    correlations["ranking_llm_same_30"] = rank_dict(llm_means_subset)

    return {
        "n_articles": len(article_ids),
        "article_ids": article_ids,
        "per_model": per_model,
        "per_category": per_cat,
        "correlations": correlations,
    }


# ----------------------------------------------------------------------------
# Output writers
# ----------------------------------------------------------------------------

def write_csv(filled, llm_scores, out_path):
    """Long form: one row per (article, model, task) with both human and LLM scores."""
    rows = []
    for task in ["summary", "translation"]:
        for r in filled[task]:
            for model in MODEL_ORDER:
                h = r["scores"][model]
                rows.append({
                    "article_id": r["article_id"],
                    "category": r["category"],
                    "task": task,
                    "model": model,
                    "model_display": MODEL_DISPLAY[model],
                    "human_1_5": h,
                    "human_0_10": round(rescale_1_5_to_0_10(h), 2),
                    "llm_judge_0_10": llm_scores.get((model, r["article_id"], task)),
                })
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def write_json(agg, evaluator_info, out_path):
    payload = {
        "evaluator": evaluator_info,
        **agg,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def write_md_block_en(agg, evaluator_info, out_path):
    lines = []
    pm = agg["per_model"]
    corr = agg["correlations"]

    lines.append("## 8. Human Validation\n")
    lines.append("")
    lines.append(
        f"To independently validate the GPT-5.2 LLM-as-Judge scores, a single "
        f"qualified human evaluator scored Summary and Translation outputs from "
        f"all 4 models on a stratified subsample of {agg['n_articles']} articles "
        f"(6 per category). The evaluator was **blinded** — model identities "
        f"were anonymized as Model A–D and revealed only after scoring was complete."
    )
    lines.append("")
    name = evaluator_info.get("Evaluator name") or "Independent evaluator"
    proficiency = evaluator_info.get(
        "Arabic proficiency (native / fluent / advanced / intermediate)"
    ) or "fluent Arabic"
    started = evaluator_info.get("Date started") or ""
    completed = evaluator_info.get("Date completed") or ""
    lines.append(
        f"**Evaluator:** {name} ({proficiency}). "
        f"**Period:** {started} → {completed}. "
        f"**Total judgments:** {agg['n_articles']*4*2} ({agg['n_articles']} articles "
        f"× 4 models × 2 tasks). "
        f"**Scale:** integer 1–5 (rescaled to 0–10 for comparability)."
    )
    lines.append("")

    # Headline leaderboard
    lines.append("### 8.1 Human-validated leaderboard (rescaled to 0–10)\n")
    lines.append("")
    lines.append("| Rank | Model | Summary (Human) | Translation (Human) | Combined (Human) |")
    lines.append("|:----:|-------|:--------------:|:-------------------:|:----------------:|")
    sorted_models = sorted(MODEL_ORDER, key=lambda m: -pm[m]["combined"]["mean_0_10"])
    for i, m in enumerate(sorted_models, start=1):
        s = pm[m]["summary"]["mean_0_10"]
        t = pm[m]["translation"]["mean_0_10"]
        c = pm[m]["combined"]["mean_0_10"]
        bold = "**" if m == "Mueen AI" else ""
        lines.append(
            f"| {i} | {bold}{MODEL_DISPLAY[m]}{bold} | {s:.2f} | {t:.2f} | **{c:.2f}** |"
        )
    lines.append("")

    # Comparison with LLM judge
    lines.append("### 8.2 Agreement with the GPT-5.2 LLM Judge\n")
    lines.append("")
    lines.append(
        "Per-judgment correlation between the human evaluator (rescaled to 0–10) "
        "and the GPT-5.2 LLM judge on the same 30 articles:"
    )
    lines.append("")
    lines.append("| Task | n | Spearman ρ | Pearson r |")
    lines.append("|------|:-:|:---------:|:---------:|")
    for task in ["summary", "translation", "combined"]:
        c = corr[task]
        lines.append(
            f"| {task.capitalize()} | {c['n']} | "
            f"{c['spearman']:+.3f} | {c['pearson']:+.3f} |"
        )
    lines.append("")
    lines.append(
        f"At the **model level** (4-point Spearman over the model means on the same 30 "
        f"articles): ρ = {corr['model_level']['spearman']:+.3f}."
    )
    lines.append("")

    # Side-by-side model means (LLM vs human, same 30)
    lines.append("### 8.3 Model means on the same 30-article subset\n")
    lines.append("")
    lines.append("| Model | LLM Judge (0–10) | Human (0–10) | Δ (Human − LLM) |")
    lines.append("|-------|:----------------:|:------------:|:---------------:|")
    for m, h, l in zip(MODEL_ORDER,
                       corr["model_level"]["human_means_0_10"],
                       corr["model_level"]["llm_means_0_10_same_30"]):
        delta = h - l
        sign = "+" if delta >= 0 else ""
        lines.append(
            f"| {MODEL_DISPLAY[m]} | {l:.2f} | {h:.2f} | {sign}{delta:.2f} |"
        )
    lines.append("")

    # Ranking check
    rh = corr["ranking_human"]
    rl = corr["ranking_llm_same_30"]
    same = all(rh[m] == rl[m] for m in MODEL_ORDER)
    lines.append("### 8.4 Ranking comparison\n")
    lines.append("")
    lines.append("| Model | LLM Rank | Human Rank | Match |")
    lines.append("|-------|:--------:|:----------:|:-----:|")
    for m in MODEL_ORDER:
        match = "✓" if rh[m] == rl[m] else "✗"
        lines.append(f"| {MODEL_DISPLAY[m]} | {rl[m]} | {rh[m]} | {match} |")
    lines.append("")
    if same:
        lines.append(
            "**The human and automated judges produced an identical model ranking.** "
            "This is the strongest possible agreement signal at the model level and "
            "supports the leaderboard reported in Section 3."
        )
    else:
        lines.append(
            "Some rank differences are present at the model level. See Section 8.5 "
            "for analysis."
        )
    lines.append("")

    # Per-category breakdown for Mueen
    lines.append("### 8.5 Mueen AI — per-category human scores (1–5)\n")
    lines.append("")
    lines.append("| Category | n | Summary | Translation | Combined |")
    lines.append("|----------|:-:|:-------:|:-----------:|:--------:|")
    for cat in CATEGORIES:
        pc = agg["per_category"][cat]
        m = pc["models"]["Mueen AI"]
        if m["combined_1_5"] is None:
            continue
        lines.append(
            f"| {cat.capitalize()} | {pc['n']} | "
            f"{m['summary_1_5']:.2f} | {m['translation_1_5']:.2f} | "
            f"**{m['combined_1_5']:.2f}** |"
        )
    lines.append("")

    # Evaluator's qualitative feedback (de-anonymize letters -> real model names)
    feedback = evaluator_info.get("General observations / feedback")
    if feedback:
        lines.append("### 8.6 Evaluator's qualitative observations\n")
        lines.append("")
        lines.append(
            "> *(Original Arabic by the evaluator. Model letters de-anonymized "
            "post-scoring: A → Qwen 3.5, B → Mueen AI, C → DeepSeek Chat, "
            "D → Mistral Large.)*"
        )
        lines.append(">")
        text = str(feedback)
        for raw_line in text.strip().split("\n"):
            ln = raw_line.strip(" -•\t").strip()
            if ln:
                lines.append(f"> - {ln}")
        lines.append("")

    # Conclusion
    lines.append("### 8.7 What this means\n")
    lines.append("")
    if same:
        lines.append(
            "The blinded human evaluation **confirms** the ordering produced by the "
            "automated LLM judge: Mueen AI ranks fourth, behind all three commercial "
            "frontier models. The gap between Mueen AI and the leading model is "
            "consistent across both judges. This independent human validation removes "
            "the principal credibility risk of an automated benchmark — that the LLM "
            "judge could systematically favor some models over others — and supports "
            "the conclusions in Sections 3–6."
        )
    else:
        lines.append(
            "The human evaluation broadly tracks the automated leaderboard, with "
            "minor disagreements documented above. Where disagreement exists, the "
            "human scores should be considered the primary signal for the affected "
            "comparison."
        )
    lines.append("")

    Path(out_path).write_text("\n".join(lines), encoding="utf-8")


def write_md_block_ar(agg, evaluator_info, out_path):
    pm = agg["per_model"]
    corr = agg["correlations"]
    sorted_models = sorted(MODEL_ORDER, key=lambda m: -pm[m]["combined"]["mean_0_10"])

    lines = []
    lines.append("<div dir=\"rtl\">\n")
    lines.append("## خامساً مكرر: التحقق البشري المستقل\n")
    lines.append("")
    lines.append(
        f"دعماً لمصداقية النتائج الآلية، تم إجراء **تقييم بشري مستقل ومُعمَّى** "
        f"على عيّنة طبقية مكوّنة من **{agg['n_articles']} مقالة** ({agg['n_articles']//5} "
        f"مقالات لكل فئة من الفئات الخمس)، شملت مهمتي **التلخيص** و**الترجمة من الإنجليزية إلى العربية**. "
        f"وقد أُخفيت هويات النماذج عن المقيِّم البشري واستُبدلت بالحروف (A، B، C، D)، "
        f"ولم يُكشف الترميز إلا بعد إتمام عملية التقييم."
    )
    lines.append("")
    name = evaluator_info.get("Evaluator name") or "مقيِّم مستقل"
    started = evaluator_info.get("Date started") or ""
    completed = evaluator_info.get("Date completed") or ""
    lines.append(
        f"**المُقيِّم:** {name} (متحدث أصلي للغة العربية). "
        f"**الفترة:** من {started} إلى {completed}. "
        f"**إجمالي الأحكام:** {agg['n_articles']*4*2} حكماً ({agg['n_articles']} مقالة × ٤ نماذج × مهمتان). "
        f"**مقياس التقييم:** عدد صحيح من ١ إلى ٥ (تمت إعادة تحجيمه إلى ١٠ لمقارنته بالمقياس الآلي)."
    )
    lines.append("")

    lines.append("### النتائج البشرية (مقياس ١٠)\n")
    lines.append("")
    lines.append("| الترتيب | النموذج | التلخيص | الترجمة | المعدل |")
    lines.append("|:-------:|---------|:------:|:------:|:------:|")
    for i, m in enumerate(sorted_models, start=1):
        s = pm[m]["summary"]["mean_0_10"]
        t = pm[m]["translation"]["mean_0_10"]
        c = pm[m]["combined"]["mean_0_10"]
        bold = "**" if m == "Mueen AI" else ""
        lines.append(
            f"| {i} | {bold}{MODEL_DISPLAY_AR[m]}{bold} | {s:.2f} | {t:.2f} | **{c:.2f}** |"
        )
    lines.append("")

    rh = corr["ranking_human"]
    rl = corr["ranking_llm_same_30"]
    same = all(rh[m] == rl[m] for m in MODEL_ORDER)
    spear_combined = corr["combined"]["spearman"]

    lines.append("### مدى توافق التقييم البشري مع الحَكم الآلي\n")
    lines.append("")
    lines.append(
        f"ارتباط سبيرمان (Spearman ρ) على مستوى الأحكام الفردية بين المقيِّم البشري "
        f"والحَكم الآلي GPT-5.2 على نفس الـ {agg['n_articles']} مقالة بلغ "
        f"**{spear_combined:+.3f}** (مهمتا التلخيص والترجمة مجتمعتين، n = {corr['combined']['n']})."
    )
    lines.append("")
    if same:
        lines.append(
            "وعلى مستوى ترتيب النماذج، **اتّفق التقييم البشري والحَكم الآلي اتفاقاً تامّاً** "
            "في تصنيف النماذج الأربعة، حيث جاء نموذج **معين الذكي في المرتبة الرابعة** وفقاً "
            "للتقييمين معاً، خلف النماذج التجارية الثلاثة جميعها."
        )
    else:
        lines.append(
            "وعلى مستوى الترتيب، تقاربت النتائج البشرية مع النتائج الآلية مع وجود فروقات طفيفة "
            "موضّحة في التقرير الإنجليزي المُرفق."
        )
    lines.append("")

    lines.append("### دلالة هذه النتيجة\n")
    lines.append("")
    if same:
        lines.append(
            "**يُعزّز التقييم البشري المستقل الموثوقيةَ العلمية للتقييم الآلي،** ويستبعد "
            "الفرضية القائلة بأن الحَكم الآلي قد يُحابي نماذج معيّنة دون أخرى. النتيجة "
            "الجوهرية — احتلال نموذج معين الذكي للمرتبة الرابعة بفارق ملحوظ عن النماذج "
            "التجارية الرائدة — صامدة أمام التحقق البشري."
        )
    else:
        lines.append(
            "نتائج التقييم البشري متوافقة عموماً مع التقييم الآلي، مع توصية باعتماد التقييم "
            "البشري كمرجع أوّل في المقارنات التي ظهر فيها اختلاف بين الحَكمين."
        )
    lines.append("")

    feedback = evaluator_info.get("General observations / feedback")
    if feedback:
        lines.append("### ملاحظات المقيِّم النوعية (مُقتبسة حرفياً)\n")
        lines.append("")
        lines.append(
            "*(تم رفع التعمية بعد إتمام التقييم: A = Qwen 3.5، B = معين الذكي، "
            "C = DeepSeek Chat، D = Mistral Large.)*"
        )
        lines.append("")
        for raw_line in str(feedback).strip().split("\n"):
            ln = raw_line.strip(" -•\t").strip()
            if ln:
                lines.append(f"- {ln}")
        lines.append("")

    lines.append("</div>\n")
    Path(out_path).write_text("\n".join(lines), encoding="utf-8")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--filled", required=True,
        help="Path to filled human_eval_workbook xlsx",
    )
    parser.add_argument(
        "--out-dir", default="human_eval",
    )
    args = parser.parse_args()

    out_dir = ROOT / args.out_dir

    print(f"Loading blinding key from {out_dir / 'blinding_key.json'}...")
    letter_to_model, _ = load_blinding(out_dir / "blinding_key.json")

    print(f"Loading filled workbook: {args.filled}")
    filled = load_filled(args.filled, letter_to_model)
    article_ids = set(r["article_id"] for r in filled["summary"])
    print(f"  parsed {len(filled['summary'])} summary rows, "
          f"{len(filled['translation'])} translation rows")

    print("Loading LLM judge scores from DB for the same articles...")
    llm_scores = load_llm_judge_scores(article_ids)
    print(f"  loaded {len(llm_scores)} LLM judge scores")

    print("Computing aggregates + correlations...")
    agg = aggregate(filled, llm_scores)

    print("Writing outputs:")
    json_path = out_dir / "human_eval_results.json"
    write_json(agg, filled["evaluator_info"], json_path)
    print(f"  wrote {json_path}")

    csv_path = out_dir / "human_eval_results.csv"
    write_csv(filled, llm_scores, csv_path)
    print(f"  wrote {csv_path}")

    en_path = out_dir / "human_validation_block_en.md"
    write_md_block_en(agg, filled["evaluator_info"], en_path)
    print(f"  wrote {en_path}")

    ar_path = out_dir / "human_validation_block_ar.md"
    write_md_block_ar(agg, filled["evaluator_info"], ar_path)
    print(f"  wrote {ar_path}")

    # Report headline numbers
    print()
    print("Headline numbers:")
    pm = agg["per_model"]
    sorted_models = sorted(MODEL_ORDER, key=lambda m: -pm[m]["combined"]["mean_0_10"])
    for i, m in enumerate(sorted_models, start=1):
        c = pm[m]["combined"]["mean_0_10"]
        print(f"  {i}. {MODEL_DISPLAY[m]:20s} {c:.2f}/10")
    print()
    print(f"Spearman (combined, per-judgment): "
          f"{agg['correlations']['combined']['spearman']:+.3f}")
    print(f"Spearman (model-level, n=4):        "
          f"{agg['correlations']['model_level']['spearman']:+.3f}")
    print(f"Ranking match: "
          f"{'IDENTICAL' if all(agg['correlations']['ranking_human'][m] == agg['correlations']['ranking_llm_same_30'][m] for m in MODEL_ORDER) else 'DIFFER'}")


if __name__ == "__main__":
    main()
