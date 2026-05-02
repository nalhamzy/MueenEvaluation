"""Build the human evaluation package.

Produces a blinded Excel workbook that a single human evaluator can use to
score Summary and Translation outputs from all 4 models on 30 stratified
articles (6 per category). The evaluator does not know which column maps to
which model — that mapping lives in a private JSON key.

Outputs (in `human_eval/`):
  1. human_eval_workbook.xlsx — hand this to the evaluator
  2. blinding_key.json        — PRIVATE: model -> column letter mapping
  3. sampled_articles.json    — list of the 30 article IDs used

Run from project root:
    py scripts/build_human_eval_package.py [--seed 7] [--per-category 6]
"""

import argparse
import json
import random
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "backend"))

from database import SessionLocal
from models import (
    Article, DatasetItem, EvaluationRun, ModelOutput,
    RunStatus, OutputStatus,
)


CATEGORIES = ["culture", "finance", "politics", "sports", "tech"]
MODELS = [
    "qwen3.5-397b-a17b",
    "deepseek-chat",
    "mistral.mistral-large-2402-v1:0",
    "Mueen AI",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SCORE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(border_style="thin", color="888888")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def gather_data():
    """Pull latest ModelOutput per (model, article) + dataset references."""
    db = SessionLocal()
    try:
        runs = db.query(EvaluationRun).filter(
            EvaluationRun.status == RunStatus.COMPLETED
        ).all()

        model_outputs = defaultdict(dict)
        for r in runs:
            outs = db.query(ModelOutput).filter(
                ModelOutput.run_id == r.id,
                ModelOutput.status == OutputStatus.SCORED,
            ).all()
            for o in outs:
                if o.article_id.startswith("ART_"):
                    continue
                prev = model_outputs[r.model_name].get(o.article_id)
                if prev is None or (
                    o.processed_at and prev.processed_at
                    and o.processed_at > prev.processed_at
                ):
                    model_outputs[r.model_name][o.article_id] = o

        articles = {
            a.id: {
                "id": a.id,
                "title": a.title,
                "body": a.body,
                "category": (a.source or a.id.split("_")[0]).lower(),
            }
            for a in db.query(Article).all()
        }

        items = {it.article_id: it for it in db.query(DatasetItem).all()}
        return {
            "model_outputs": dict(model_outputs),
            "articles": articles,
            "items": items,
        }
    finally:
        db.close()


def pick_articles(data, per_category: int, seed: int):
    """Stratified random sample: `per_category` articles per category.

    Only consider articles where ALL 4 models produced a scored output for
    BOTH summary and translation tasks (so the evaluator sees no blanks).
    """
    rng = random.Random(seed)
    eligible_by_cat = defaultdict(list)

    all_article_ids = set(data["articles"].keys())
    for aid in all_article_ids:
        item = data["items"].get(aid)
        if not item or not item.summary_reference or not item.translation_input:
            continue
        ok = True
        for m in MODELS:
            o = data["model_outputs"].get(m, {}).get(aid)
            if not o or not (o.summary_output and o.translation_output):
                ok = False
                break
        if not ok:
            continue
        cat = data["articles"][aid]["category"]
        if cat in CATEGORIES:
            eligible_by_cat[cat].append(aid)

    picked = []
    for cat in CATEGORIES:
        pool = sorted(eligible_by_cat[cat])
        rng.shuffle(pool)
        picked.extend(pool[:per_category])
    return picked


def build_blinding(seed: int):
    """Map each real model to a letter A..D. Deterministic per seed."""
    rng = random.Random(seed + 1)
    letters = ["A", "B", "C", "D"]
    rng.shuffle(letters)
    return dict(zip(MODELS, letters))


# ----------------------------------------------------------------------------
# Sheet builders
# ----------------------------------------------------------------------------

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row, values):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def build_instructions(ws):
    ws.title = "Instructions"
    _set_col_widths(ws, [110])
    ws.row_dimensions[1].height = 30
    cell = ws.cell(row=1, column=1, value="Human Evaluation — Arabic LLM Benchmark")
    cell.font = Font(bold=True, size=16, color="1F4E78")

    blocks = [
        ("Purpose",
         "You are helping validate an automated benchmark of Arabic language models. "
         "Four anonymous models (labeled Model A, B, C, D) produced Arabic summaries "
         "and English→Arabic translations. Your judgment provides an independent check "
         "on the automated scoring."),

        ("Important — blinding",
         "You will NOT be told which model is which. Please do not attempt to guess. "
         "The same letter (A/B/C/D) refers to the same model consistently across all "
         "rows in both task sheets."),

        ("What to do",
         "1. Open the 'Summary Task' sheet. For each row:\n"
         "   a. Read the Arabic article (col. Article).\n"
         "   b. Read the reference summary (col. Reference Summary).\n"
         "   c. Read each model's summary (cols. Model A–D).\n"
         "   d. Give each model a score from 1 to 5 (see scale below).\n"
         "   e. Optionally add a short note explaining low/high scores.\n"
         "2. Move to the 'Translation Task' sheet and repeat.\n"
         "3. Fill in the 'Evaluator Info' sheet when done.\n"
         "4. Save the file and return it to the requester."),

        ("Scoring scale (1–5, integer only)",
         "5 = Excellent — publication quality, faithful and fluent, no meaningful issues.\n"
         "4 = Good — faithful and fluent, minor issues only (small awkwardness, minor omission).\n"
         "3 = Acceptable — main content correct but clear issues (coverage gaps, some awkward phrasing).\n"
         "2 = Poor — significant errors (notable factual issues, poor Arabic, missing key content).\n"
         "1 = Unacceptable — factually wrong, unreadable, or unusable as Arabic output."),

        ("What counts as a factual error (Summary)",
         "Anything the summary claims that is NOT supported by the article body. "
         "Extra details the model invented = factual error. "
         "Missing a detail is a coverage issue, not a factual error."),

        ("What counts as a faithfulness error (Translation)",
         "Anything in the Arabic translation that is not present in the English source, "
         "or any English meaning dropped in the Arabic. Translations should be formal MSA "
         "(فصحى) — noticeable dialect use is a fluency/register issue."),

        ("Do",
         "• Score each row independently — don't let earlier rows anchor you.\n"
         "• Take breaks. This sheet is expected to take ~6 hours total.\n"
         "• If an Arabic output is cut off or empty, score 1 and note it.\n"
         "• Score all 4 models on every row even if they look similar."),

        ("Do NOT",
         "• Do not try to identify which letter maps to which model.\n"
         "• Do not discuss scores with anyone else while working.\n"
         "• Do not use any AI tool or translator to help you judge.\n"
         "• Do not leave score cells blank — every row needs 4 integers."),

        ("Dataset summary",
         "30 articles total: 6 from each of 5 categories (culture, finance, politics, "
         "sports, tech). Each article is ~200–400 words of Arabic. The same article "
         "appears in both the Summary sheet and the Translation sheet with the same "
         "article_id, but the tasks are independent — score them on their own merits."),

        ("Contact",
         "If you find an obviously corrupted row (e.g., all models produced empty output), "
         "enter score 1 across the row and add a note. Return the workbook by the "
         "agreed deadline."),
    ]

    row = 3
    for title, body in blocks:
        t = ws.cell(row=row, column=1, value=title)
        t.font = Font(bold=True, size=12, color="1F4E78")
        row += 1
        b = ws.cell(row=row, column=1, value=body)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(30, 15 * (body.count("\n") + body.count(". ") // 2 + 2))
        row += 2


def build_task_sheet(ws, title, rows_data, task_name):
    """rows_data: list of dicts with keys:
         article_id, category, context_text, reference_text,
         model_A, model_B, model_C, model_D
    task_name: 'Summary' or 'Translation' (for sheet header)
    """
    ws.title = title

    headers = [
        "Row", "Article ID", "Category",
        "Article" if task_name == "Summary" else "English Source",
        "Reference Arabic",
        "Model A", "Model B", "Model C", "Model D",
        "Score A (1-5)", "Score B (1-5)", "Score C (1-5)", "Score D (1-5)",
        "Notes (optional)",
    ]
    widths = [5, 14, 10, 55, 40, 40, 40, 40, 40, 10, 10, 10, 10, 40]
    _set_col_widths(ws, widths)
    _header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "F2"

    # Score validation: integers 1–5
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=False,
        errorTitle="Invalid score",
        error="Score must be an integer 1–5",
        promptTitle="Score",
        prompt="Enter an integer between 1 and 5",
    )
    dv.showErrorMessage = True
    ws.add_data_validation(dv)

    for i, r in enumerate(rows_data, start=1):
        xr = i + 1
        ws.cell(row=xr, column=1, value=i)
        ws.cell(row=xr, column=2, value=r["article_id"])
        ws.cell(row=xr, column=3, value=r["category"])
        ws.cell(row=xr, column=4, value=r["context_text"])
        ws.cell(row=xr, column=5, value=r["reference_text"])
        ws.cell(row=xr, column=6, value=r["model_A"])
        ws.cell(row=xr, column=7, value=r["model_B"])
        ws.cell(row=xr, column=8, value=r["model_C"])
        ws.cell(row=xr, column=9, value=r["model_D"])
        # Score cells
        for col in range(10, 14):
            c = ws.cell(row=xr, column=col)
            c.fill = SCORE_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BOX
        # Wrap text in long cells
        for col in [4, 5, 6, 7, 8, 9, 14]:
            c = ws.cell(row=xr, column=col)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if task_name == "Summary" and col == 4:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[xr].height = 130

        dv.add(f"{get_column_letter(10)}{xr}:{get_column_letter(13)}{xr}")


def build_evaluator_info(ws):
    ws.title = "Evaluator Info"
    _set_col_widths(ws, [35, 60])
    fields = [
        ("Evaluator name", ""),
        ("Date started", ""),
        ("Date completed", ""),
        ("Total hours spent", ""),
        ("Arabic proficiency (native / fluent / advanced / intermediate)", ""),
        ("Confidence in your scoring (1-5)", ""),
        ("General observations / feedback", ""),
    ]
    for i, (label, _) in enumerate(fields, start=1):
        lbl = ws.cell(row=i, column=1, value=label)
        lbl.font = Font(bold=True)
        lbl.fill = SUBHDR_FILL
        lbl.border = BOX
        val = ws.cell(row=i, column=2, value="")
        val.border = BOX
        val.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30
    ws.row_dimensions[len(fields)].height = 80


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--seed", type=int, default=7)
    parser.add_argument("--per-category", type=int, default=6)
    parser.add_argument("--out-dir", default="human_eval")
    args = parser.parse_args()

    out_dir = ROOT / args.out_dir
    out_dir.mkdir(exist_ok=True)

    print("Gathering data from DB...")
    data = gather_data()

    print(f"Sampling {args.per_category} articles per category (seed={args.seed})...")
    article_ids = pick_articles(data, args.per_category, args.seed)
    print(f"  selected {len(article_ids)} articles")

    blinding = build_blinding(args.seed)
    # inverse map for building rows: letter -> model_key
    letter_to_model = {v: k for k, v in blinding.items()}

    # Build row payloads for both tasks
    summary_rows = []
    translation_rows = []
    for aid in article_ids:
        art = data["articles"][aid]
        item = data["items"][aid]

        row_base_sum = {
            "article_id": aid,
            "category": art["category"],
            "context_text": art["body"],
            "reference_text": item.summary_reference or "",
        }
        row_base_tx = {
            "article_id": aid,
            "category": art["category"],
            "context_text": item.translation_input or "",
            "reference_text": item.translation_reference or "",
        }
        for letter in ["A", "B", "C", "D"]:
            model_key = letter_to_model[letter]
            o = data["model_outputs"][model_key][aid]
            row_base_sum[f"model_{letter}"] = o.summary_output or ""
            row_base_tx[f"model_{letter}"] = o.translation_output or ""
        summary_rows.append(row_base_sum)
        translation_rows.append(row_base_tx)

    print("Building workbook...")
    wb = Workbook()
    build_instructions(wb.active)
    build_task_sheet(wb.create_sheet(), "Summary Task", summary_rows, "Summary")
    build_task_sheet(wb.create_sheet(), "Translation Task", translation_rows, "Translation")
    build_evaluator_info(wb.create_sheet())

    xlsx_path = out_dir / "human_eval_workbook.xlsx"
    wb.save(xlsx_path)
    print(f"  wrote {xlsx_path}")

    # Private blinding key (not for evaluator!)
    key_path = out_dir / "blinding_key.json"
    key_payload = {
        "seed": args.seed,
        "per_category": args.per_category,
        "total_articles": len(article_ids),
        "model_to_letter": blinding,
        "letter_to_model": letter_to_model,
        "warning": "PRIVATE — do not share with the evaluator until scoring is complete.",
    }
    with open(key_path, "w", encoding="utf-8") as f:
        json.dump(key_payload, f, indent=2, ensure_ascii=False)
    print(f"  wrote {key_path}")

    # Sampled article IDs (public)
    sampled_path = out_dir / "sampled_articles.json"
    with open(sampled_path, "w", encoding="utf-8") as f:
        json.dump({
            "seed": args.seed,
            "per_category": args.per_category,
            "article_ids": article_ids,
            "by_category": {
                cat: [a for a in article_ids if data["articles"][a]["category"] == cat]
                for cat in CATEGORIES
            },
        }, f, indent=2, ensure_ascii=False)
    print(f"  wrote {sampled_path}")

    print("\nBlinding (keep private):")
    for m, l in blinding.items():
        print(f"  {m:40s} -> Model {l}")


if __name__ == "__main__":
    main()
